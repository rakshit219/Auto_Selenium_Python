import openpyxl

def getRowCount(path,sheetname):
    wb=openpyxl.load_workbook(path)
    sheet=wb[sheetname]
    return(sheet.max_row)

def getColumnCount(path,sheetname):
    wb=openpyxl.load_workbook(path)
    sheet=wb[sheetname]
    return(sheet.max_column)

def readData(path,sheetname,rownum,colnum):
    wb=openpyxl.load_workbook(path)
    sheet=wb[sheetname]
    return sheet.cell(row=rownum,column=colnum).value

def writeData(path,sheetname,rownum,colnum,data):
    wb=openpyxl.load_workbook(path)
    sheet=wb[sheetname]
    sheet.cell(row=rownum,column=colnum).value=data
    wb.save(path)
